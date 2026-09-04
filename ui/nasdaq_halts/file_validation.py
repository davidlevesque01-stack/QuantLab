from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
import csv


@dataclass
class FileValidationResult:
    valid: bool
    file_name: str
    file_format: str
    observation_count: int
    date_column_found: bool
    ticker_column_found: bool
    invalid_dates: int
    empty_tickers: int
    errors: list[str]


def validate_input_file(path: str, separator: str = ",") -> FileValidationResult:
    file_path = Path(path)

    if not file_path.exists():
        return _failure(file_path, "Input file does not exist.")

    suffix = file_path.suffix.lower()

    if suffix == ".xlsx":
        return _validate_xlsx(file_path)

    if suffix == ".csv":
        return _validate_csv(file_path, separator)

    return _failure(
        file_path,
        "Unsupported file format. Use XLSX or CSV.",
        "Unknown",
    )


def _failure(
    file_path: Path,
    message: str,
    file_format: str = "Unknown",
) -> FileValidationResult:
    return FileValidationResult(
        False, file_path.name, file_format, 0,
        False, False, 0, 0, [message]
    )


def _validate_csv(file_path: Path, separator: str) -> FileValidationResult:
    try:
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=separator)
            fieldnames = reader.fieldnames or []
            normalized = {
                name.strip().lower(): name
                for name in fieldnames
                if name is not None
            }

            date_column = normalized.get("date")
            ticker_column = normalized.get("ticker")

            if not date_column or not ticker_column:
                missing = []
                if not date_column:
                    missing.append("Date")
                if not ticker_column:
                    missing.append("Ticker")
                return FileValidationResult(
                    False, file_path.name, "CSV", 0,
                    bool(date_column), bool(ticker_column), 0, 0,
                    [f"Missing required column(s): {', '.join(missing)}."]
                )

            count = 0
            invalid_dates = 0
            empty_tickers = 0

            for row in reader:
                count += 1
                ticker = (row.get(ticker_column) or "").strip()
                raw_date = (row.get(date_column) or "").strip()

                if not ticker:
                    empty_tickers += 1

                if not _valid_date(raw_date):
                    invalid_dates += 1

            valid = invalid_dates == 0 and empty_tickers == 0

            return FileValidationResult(
                valid, file_path.name, "CSV", count,
                True, True, invalid_dates, empty_tickers,
                [] if valid else ["One or more observations are invalid."]
            )

    except UnicodeDecodeError:
        return _failure(file_path, "Unable to read the CSV file as UTF-8.", "CSV")
    except csv.Error as exc:
        return _failure(file_path, f"CSV parsing error: {exc}", "CSV")
    except OSError as exc:
        return _failure(file_path, f"Unable to read the input file: {exc}", "CSV")


def _validate_xlsx(file_path: Path) -> FileValidationResult:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return _failure(
            file_path,
            "openpyxl is required to validate XLSX files.",
            "XLSX",
        )

    try:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)

        try:
            header = next(rows)
        except StopIteration:
            workbook.close()
            return _failure(file_path, "The XLSX file is empty.", "XLSX")

        fieldnames = [
            str(value).strip() if value is not None else ""
            for value in header
        ]
        normalized = {
            name.lower(): index
            for index, name in enumerate(fieldnames)
            if name
        }

        date_index = normalized.get("date")
        ticker_index = normalized.get("ticker")

        if date_index is None or ticker_index is None:
            missing = []
            if date_index is None:
                missing.append("Date")
            if ticker_index is None:
                missing.append("Ticker")
            workbook.close()
            return FileValidationResult(
                False, file_path.name, "XLSX", 0,
                date_index is not None, ticker_index is not None,
                0, 0,
                [f"Missing required column(s): {', '.join(missing)}."]
            )

        count = 0
        invalid_dates = 0
        empty_tickers = 0

        for row in rows:
            count += 1
            ticker_value = row[ticker_index] if ticker_index < len(row) else None
            date_value = row[date_index] if date_index < len(row) else None

            if ticker_value is None or not str(ticker_value).strip():
                empty_tickers += 1

            if not _valid_date(date_value):
                invalid_dates += 1

        workbook.close()

        valid = invalid_dates == 0 and empty_tickers == 0

        return FileValidationResult(
            valid, file_path.name, "XLSX", count,
            True, True, invalid_dates, empty_tickers,
            [] if valid else ["One or more observations are invalid."]
        )

    except Exception as exc:
        return _failure(
            file_path,
            f"Unable to read the XLSX file: {exc}",
            "XLSX",
        )


def _valid_date(value) -> bool:
    if isinstance(value, datetime):
        return True

    if value is None:
        return False

    text = str(value).strip()

    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            datetime.strptime(text, fmt)
            return True
        except ValueError:
            pass

    return False
